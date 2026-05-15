import csv
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_dt(s):
    s = s.strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%d %H:%M:%S.%f +0000 UTC', '%Y-%m-%d %H:%M:%S +0000 UTC'):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


with open(r'C:\Users\PrashantKumar\Downloads\1778877418890839804.csv', encoding='utf-8-sig') as f:
    rows = [r for r in csv.DictReader(f) if r['Severity'].strip() == 'HIGH']

snaps = [
    ('Mar 15', datetime(2026, 3, 15, 23, 59, 59)),
    ('Apr 15', datetime(2026, 4, 15, 23, 59, 59)),
    ('May 15', datetime(2026, 5, 15, 23, 59, 59)),
]

# Compute snapshot counts
def open_at(snap_date):
    return sum(
        1 for r in rows
        if (c := parse_dt(r['Created At'])) and c <= snap_date
        and ((rv := parse_dt(r['Resolved Time'])) is None or rv > snap_date)
    )

def resolved_by(snap_date, status_filter=None):
    return sum(
        1 for r in rows
        if (not status_filter or r['Status'].strip() == status_filter)
        and r['Status'].strip() in ('RESOLVED', 'REJECTED')
        and (rv := parse_dt(r['Resolved Time'])) and rv <= snap_date
    )

def ignored_by(snap_date):
    return sum(
        1 for r in rows
        if r['Status'].strip() == 'REJECTED'
        and (rv := parse_dt(r['Resolved Time'])) and rv <= snap_date
    )

def total_by(snap_date):
    return sum(1 for r in rows if (c := parse_dt(r['Created At'])) and c <= snap_date)

mar_sd, apr_sd, may_sd = datetime(2026,3,15,23,59,59), datetime(2026,4,15,23,59,59), datetime(2026,5,15,23,59,59)

mar_o, apr_o, may_o = open_at(mar_sd), open_at(apr_sd), open_at(may_sd)
mar_t, apr_t, may_t = total_by(mar_sd), total_by(apr_sd), total_by(may_sd)
mar_ri, apr_ri, may_ri = resolved_by(mar_sd), resolved_by(apr_sd), resolved_by(may_sd)
mar_ig, apr_ig, may_ig = ignored_by(mar_sd), ignored_by(apr_sd), ignored_by(may_sd)
mar_res = mar_ri - mar_ig
apr_res = apr_ri - apr_ig
may_res = may_ri - may_ig

res_all = sum(1 for r in rows if r['Status'].strip() == 'RESOLVED')
ign_all = sum(1 for r in rows if r['Status'].strip() == 'REJECTED')

def pct(n, d): return f'{n/d*100:.1f}%' if d else '0%'
def pdiff(a, b): s = '+' if b-a >= 0 else ''; return f'{s}{b-a:.1f}%'
def cdiff(a, b):
    d = b - a
    s = '+' if d >= 0 else ''
    return f'{s}{d}'

mar_pct_val = mar_ri / mar_t * 100 if mar_t else 0
apr_pct_val = apr_ri / apr_t * 100 if apr_t else 0
may_pct_val = may_ri / may_t * 100 if may_t else 0

metrics = [
    ('Total Issues Created', mar_t, apr_t, may_t,
     f'+{may_t-mar_t}', f'+{apr_t-mar_t}', f'+{may_t-apr_t}',
     PatternFill('solid', fgColor='F2F2F2')),
    ('Open / In Progress', mar_o, apr_o, may_o,
     cdiff(mar_o, may_o), cdiff(mar_o, apr_o), cdiff(apr_o, may_o),
     PatternFill('solid', fgColor='FCE4D6')),
    ('Resolved  (Cumulative)', mar_res, apr_res, may_res,
     f'+{may_res-mar_res}', f'+{apr_res-mar_res}', f'+{may_res-apr_res}',
     None),
    ('Ignored / Accepted Risk  (Cumulative)', mar_ig, apr_ig, may_ig,
     f'+{may_ig-mar_ig}', f'+{apr_ig-mar_ig}', f'+{may_ig-apr_ig}',
     None),
    ('Resolution %  (Resolved + Ignored)',
     pct(mar_ri, mar_t), pct(apr_ri, apr_t), pct(may_ri, may_t),
     pdiff(mar_pct_val, may_pct_val),
     pdiff(mar_pct_val, apr_pct_val),
     pdiff(apr_pct_val, may_pct_val),
     PatternFill('solid', fgColor='FFF2CC')),
]

# Resolved/Ignored issues sorted by resolved time
ri_issues = [r for r in rows if r['Status'].strip() in ('RESOLVED', 'REJECTED')]
ri_issues.sort(key=lambda r: parse_dt(r['Resolved Time']) or datetime.max)

def snap_month(dt):
    if dt is None: return '-'
    if dt <= mar_sd: return 'March 2026'
    if dt <= apr_sd: return 'April 2026'
    if dt <= may_sd: return 'May 2026'
    return dt.strftime('%b %Y')

# ── Styles ────────────────────────────────────────────────────────────────────
thin = Side(style='thin', color='BDD7EE')
def cell_border(): return Border(top=thin, bottom=thin, left=thin, right=thin)

HDR_DARK  = PatternFill('solid', fgColor='1F4E79')
HDR_GREEN = PatternFill('solid', fgColor='375623')
ORG_FILL  = PatternFill('solid', fgColor='C55A11')
ZEBRA     = PatternFill('solid', fgColor='FFF2EC')
CTR_WRAP  = Alignment(horizontal='center', vertical='center', wrap_text=True)
CTR       = Alignment(horizontal='center', vertical='center')
LEFT_PAD  = Alignment(horizontal='left', vertical='center', indent=1)
WHITE_BOLD = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
DARK_BOLD  = Font(name='Calibri', bold=True, size=11, color='1F4E79')
REG        = Font(name='Calibri', size=11)

def h(ws, row, col, val, fill=None, font=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=val)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c

# ── Build tab ─────────────────────────────────────────────────────────────────
wb = load_workbook(r'C:\Users\PrashantKumar\csv2excel\CareFirst_Governance_Report.xlsx')

# Remove old HIGH tab if exists
if 'HIGH Issues' in wb.sheetnames:
    del wb['HIGH Issues']

ws = wb.create_sheet('HIGH Issues')

# Title
ws.merge_cells('A1:G1')
h(ws, 1, 1, 'HIGH Severity — Resolution Progress (Month-on-Month)',
  fill=ORG_FILL, font=Font(name='Calibri', bold=True, size=14, color='FFFFFF'), align=CTR_WRAP)
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:G2')
h(ws, 2, 1, 'Exec Summary  |  Severity: HIGH  |  As of May 15, 2026',
  fill=PatternFill('solid', fgColor='3C3C3C'),
  font=Font(name='Calibri', italic=True, size=10, color='FFFFFF'), align=LEFT_PAD)
ws.row_dimensions[2].height = 18

# Section labels
ws.merge_cells('A3:D3')
h(ws, 3, 1, 'SNAPSHOT DATA', fill=PatternFill('solid', fgColor='2E75B6'), font=WHITE_BOLD, align=CTR)
for ci in range(2, 5):
    ws.cell(row=3, column=ci).fill = PatternFill('solid', fgColor='2E75B6')
ws.merge_cells('E3:G3')
h(ws, 3, 5, 'PROGRESS TRACKING', fill=HDR_GREEN, font=WHITE_BOLD, align=CTR)
for ci in range(6, 8):
    ws.cell(row=3, column=ci).fill = HDR_GREEN
ws.row_dimensions[3].height = 22

# Column headers
col_headers = [
    'Metric', 'Mar 15, 2026\n(Baseline)', 'Apr 15, 2026', 'May 15, 2026\n(Latest)',
    'vs Baseline\n(Mar -> May)', 'Monthly\nMar -> Apr', 'Monthly\nApr -> May',
]
for ci, ch in enumerate(col_headers, 1):
    fill = HDR_DARK if ci <= 4 else HDR_GREEN
    h(ws, 4, ci, ch, fill=fill, font=WHITE_BOLD, align=CTR_WRAP, border=cell_border())
ws.row_dimensions[4].height = 40

# Metrics rows
for ri, row in enumerate(metrics, 5):
    label, mar, apr, may, c1, c2, c3, row_fill = row
    is_pct   = 'Resolution %' in str(label)
    for ci, val in enumerate([label, mar, apr, may, c1, c2, c3], 1):
        ff = row_fill if row_fill else (ZEBRA if ri % 2 == 0 else None)
        fn = Font(name='Calibri', bold=True, size=11, color='7F6000') if is_pct \
             else (DARK_BOLD if ci == 1 else REG)
        h(ws, ri, ci, val, fill=ff, font=fn,
          align=LEFT_PAD if ci == 1 else CTR, border=cell_border())
    ws.row_dimensions[ri].height = 20

# EOD footnote
footnote_row = 5 + len(metrics)
ws.merge_cells(f'A{footnote_row}:G{footnote_row}')
h(ws, footnote_row, 1,
  '* Snapshots taken as of End-of-Day (23:59:59) on each date.',
  font=Font(name='Calibri', italic=True, size=9, color='808080'),
  align=LEFT_PAD)
ws.row_dimensions[footnote_row].height = 16

# Spacer
spacer_row = footnote_row + 1
ws.row_dimensions[spacer_row].height = 14

# Issue list section header
issue_list_start = spacer_row + 1
ws.merge_cells(f'A{issue_list_start}:G{issue_list_start}')
h(ws, issue_list_start, 1, 'Resolved Issues Detail  (Resolved + Ignored)',
  fill=PatternFill('solid', fgColor='3C3C3C'),
  font=Font(name='Calibri', bold=True, size=11, color='FFFFFF'), align=LEFT_PAD)
ws.row_dimensions[issue_list_start].height = 22

ws.merge_cells(f'A{issue_list_start+1}:G{issue_list_start+1}')
h(ws, issue_list_start+1, 1, 'IGNORED = team has accepted the risk and committed to a fix-by date',
  fill=PatternFill('solid', fgColor='595959'),
  font=Font(name='Calibri', italic=True, size=9, color='FFFFFF'), align=LEFT_PAD)
ws.row_dimensions[issue_list_start+1].height = 16

# List column headers
list_hdr_row = issue_list_start + 2
list_headers = ['Issue ID', 'Title', 'Severity', 'Status', 'Created At', 'Resolved Time', 'Snapshot Month']
for ci, ch in enumerate(list_headers, 1):
    h(ws, list_hdr_row, ci, ch, fill=HDR_DARK, font=WHITE_BOLD, align=CTR_WRAP, border=cell_border())
ws.row_dimensions[list_hdr_row].height = 28

# Issues
status_colors = {'RESOLVED': '375623', 'REJECTED': 'BF8F00'}
for ri, r in enumerate(ri_issues, list_hdr_row + 1):
    row_fill = ZEBRA if ri % 2 == 0 else None
    status = r['Status'].strip()
    display_status = 'IGNORED' if status == 'REJECTED' else status
    s_fill = PatternFill('solid', fgColor=status_colors.get(status, '666666'))
    created_dt  = parse_dt(r['Created At'])
    resolved_dt = parse_dt(r['Resolved Time'])
    vals = [
        r['Issue ID'].strip(),
        r['Title'].strip(),
        'HIGH',
        display_status,
        created_dt.strftime('%Y-%m-%d') if created_dt else '',
        resolved_dt.strftime('%Y-%m-%d %H:%M') if resolved_dt else '',
        snap_month(resolved_dt),
    ]
    for ci, val in enumerate(vals, 1):
        ff = s_fill if ci == 4 else row_fill
        if ci == 3:
            fn = Font(name='Calibri', bold=True, size=10, color='C55A11')
        elif ci == 4:
            fn = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        else:
            fn = Font(name='Calibri', size=10)
        h(ws, ri, ci, val, fill=ff, font=fn,
          align=LEFT_PAD if ci in (1, 2) else CTR, border=cell_border())
    ws.row_dimensions[ri].height = 18

# Footer
footer_row = list_hdr_row + 1 + len(ri_issues)
ws.merge_cells(f'A{footer_row}:G{footer_row}')
h(ws, footer_row, 1,
  f'Total: {len(ri_issues)}  |  Resolved: {res_all}  |  Ignored: {ign_all}  |  {len(rows)-len(ri_issues)} issues still Open/In-Progress (not listed)',
  fill=HDR_DARK, font=Font(name='Calibri', bold=True, size=10, color='FFFFFF'), align=LEFT_PAD)
ws.row_dimensions[footer_row].height = 20

# Column widths
for ci, w in enumerate([38, 42, 10, 10, 14, 22, 14], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = 'A5'

wb.save(r'C:\Users\PrashantKumar\csv2excel\CareFirst_Governance_Report.xlsx')
print(f'Done — HIGH Issues tab added. {len(ri_issues)} resolved/ignored issues listed.')
