import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description="Apply print header/footer watermark to every sheet in a workbook.")
parser.add_argument("workbook", help="Path to the .xlsx workbook")
args = parser.parse_args()

wb = load_workbook(args.workbook)

header_text = '&"Calibri,Bold"&36&KA6A6A6CONFIDENTIAL \u2014 PROPRIETARY TO AHEAD'
footer_left  = '&"Calibri,Regular"&9&K808080Confidential \u2014 Proprietary to AHEAD'
footer_right = '&"Calibri,Regular"&9&K808080Page &P of &N'

for ws in wb.worksheets:
    ws.oddHeader.center.text = header_text
    ws.oddFooter.left.text   = footer_left
    ws.oddFooter.right.text  = footer_right

wb.save(args.workbook)
print('Watermark added to tabs:', wb.sheetnames)
