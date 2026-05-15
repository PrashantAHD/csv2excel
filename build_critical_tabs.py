from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook(r'C:\Users\PrashantKumar\csv2excel\CareFirst_Governance_Report.xlsx')

# ── shared styles ─────────────────────────────────────────────────────────────
thin = Side(style='thin', color='BDD7EE')

def cell_border():
    return Border(top=thin, bottom=thin, left=thin, right=thin)

HDR_DARK   = PatternFill('solid', fgColor='1F4E79')
HDR_GREEN  = PatternFill('solid', fgColor='375623')
RED_FILL   = PatternFill('solid', fgColor='C00000')
ZEBRA      = PatternFill('solid', fgColor='EBF3FB')

CTR_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
CTR      = Alignment(horizontal='center', vertical='center')
LEFT_PAD = Alignment(horizontal='left', vertical='center', indent=1)

WHITE_BOLD = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
DARK_BOLD  = Font(name='Calibri', bold=True, size=11, color='1F4E79')
REG        = Font(name='Calibri', size=11)


def h(ws, row, col, val, fill=None, font=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=val)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c


# ══════════════════════════════════════════════════════════════════════════════
# SINGLE TAB — CRITICAL Issues
# ══════════════════════════════════════════════════════════════════════════════
for sheet in wb.sheetnames:
    del wb[sheet]
ws = wb.create_sheet('CRITICAL Issues')

# ── SECTION 1: Title ──────────────────────────────────────────────────────────
ws.merge_cells('A1:G1')
h(ws, 1, 1, 'CRITICAL Severity — Resolution Progress (Month-on-Month)',
  fill=RED_FILL,
  font=Font(name='Calibri', bold=True, size=14, color='FFFFFF'),
  align=CTR_WRAP)
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:G2')
h(ws, 2, 1, 'Exec Summary  |  Severity: CRITICAL  |  As of May 15, 2026',
  fill=PatternFill('solid', fgColor='3C3C3C'),
  font=Font(name='Calibri', italic=True, size=10, color='FFFFFF'),
  align=LEFT_PAD)
ws.row_dimensions[2].height = 18

# ── SECTION 2: Section label row ─────────────────────────────────────────────
ws.merge_cells('A3:D3')
h(ws, 3, 1, 'SNAPSHOT DATA',
  fill=PatternFill('solid', fgColor='2E75B6'), font=WHITE_BOLD, align=CTR)
for ci in range(2, 5):
    ws.cell(row=3, column=ci).fill = PatternFill('solid', fgColor='2E75B6')
ws.merge_cells('E3:G3')
h(ws, 3, 5, 'PROGRESS TRACKING',
  fill=HDR_GREEN, font=WHITE_BOLD, align=CTR)
for ci in range(6, 8):
    ws.cell(row=3, column=ci).fill = HDR_GREEN
ws.row_dimensions[3].height = 22

# ── SECTION 3: Column headers ─────────────────────────────────────────────────
col_headers = [
    'Metric',
    'Mar 15, 2026\n(Baseline)',
    'Apr 15, 2026',
    'May 15, 2026\n(Latest)',
    'vs Baseline\n(Mar -> May)',
    'Monthly\nMar -> Apr',
    'Monthly\nApr -> May',
]
for ci, ch in enumerate(col_headers, 1):
    fill = HDR_DARK if ci <= 4 else HDR_GREEN
    h(ws, 4, ci, ch, fill=fill, font=WHITE_BOLD, align=CTR_WRAP, border=cell_border())
ws.row_dimensions[4].height = 40

# ── SECTION 4: Metrics rows ───────────────────────────────────────────────────
# Format: (label, mar, apr, may, vs_baseline, mar->apr change, apr->may change, row_fill)
metrics = [
    ('Total Issues Created',
     8, 8, 8,
     '-', '-', '-',
     PatternFill('solid', fgColor='F2F2F2')),
    ('Open / In Progress',
     8, 5, 4,
     '-4', '-3', '-1',
     PatternFill('solid', fgColor='FCE4D6')),
    ('Resolved  (Cumulative)',
     0, 2, 3,
     '+3', '+2', '+1',
     None),
    ('Ignored / Accepted Risk  (Cumulative)',
     0, 1, 1,
     '+1', '+1', '-',
     None),
    ('Resolution %  (Resolved + Ignored)',
     '0%', '37.5%', '50.0%',
     '+50.0%', '+37.5%', '+12.5%',
     PatternFill('solid', fgColor='FFF2CC')),
]

for ri, row in enumerate(metrics, 5):
    label, mar, apr, may, c1, c2, c3, row_fill = row
    is_pct    = 'Resolution %' in str(label)
    for ci, val in enumerate([label, mar, apr, may, c1, c2, c3], 1):
        ff = row_fill if row_fill else (ZEBRA if ri % 2 == 0 else None)
        fn = Font(name='Calibri', bold=True, size=11, color='7F6000') if is_pct \
             else (DARK_BOLD if ci == 1 else REG)
        h(ws, ri, ci, val, fill=ff, font=fn,
          align=LEFT_PAD if ci == 1 else CTR, border=cell_border())
    ws.row_dimensions[ri].height = 20

# ── EOD footnote ──────────────────────────────────────────────────────────────
footnote_row = 5 + len(metrics)
ws.merge_cells(f'A{footnote_row}:G{footnote_row}')
h(ws, footnote_row, 1,
  '* Snapshots taken as of End-of-Day (23:59:59) on each date.',
  font=Font(name='Calibri', italic=True, size=9, color='808080'),
  align=LEFT_PAD)
ws.row_dimensions[footnote_row].height = 16

# ── SPACER ROW ────────────────────────────────────────────────────────────────
spacer_row = footnote_row + 1
ws.row_dimensions[spacer_row].height = 14

# ── SECTION 5: Issue list title ───────────────────────────────────────────────
issue_start = spacer_row + 1
ws.merge_cells(f'A{issue_start}:G{issue_start}')
h(ws, issue_start, 1, 'Resolved Issues Detail  (Resolved + Ignored)',
  fill=PatternFill('solid', fgColor='3C3C3C'),
  font=Font(name='Calibri', bold=True, size=11, color='FFFFFF'),
  align=LEFT_PAD)
ws.row_dimensions[issue_start].height = 22

ws.merge_cells(f'A{issue_start+1}:G{issue_start+1}')
h(ws, issue_start+1, 1,
  'IGNORED = team has accepted the risk and committed to a fix-by date',
  fill=PatternFill('solid', fgColor='595959'),
  font=Font(name='Calibri', italic=True, size=9, color='FFFFFF'),
  align=LEFT_PAD)
ws.row_dimensions[issue_start+1].height = 16

# ── SECTION 6: Issue list headers ─────────────────────────────────────────────
list_hdr_row = issue_start + 2
list_headers = ['Issue ID', 'Title', 'Severity', 'Status', 'Created At', 'Resolved Time', 'Snapshot Month']
for ci, ch in enumerate(list_headers, 1):
    h(ws, list_hdr_row, ci, ch, fill=HDR_DARK, font=WHITE_BOLD, align=CTR_WRAP, border=cell_border())
ws.row_dimensions[list_hdr_row].height = 28

# ── SECTION 7: Issues data ─────────────────────────────────────────────────────
issues = [
    ('ba78f854-3fdb-4ecc-b045-7775fbe2ae34', 'Publicly exposed AWS resources with sensitive data',        'CRITICAL', 'RESOLVED', '2026-01-29', '2026-04-10 19:46', 'April 2026'),
    ('f0af25bd-ef33-4ad5-9fac-b791c41bc044', 'Publicly readable bucket contains sensitive data',           'CRITICAL', 'RESOLVED', '2026-01-29', '2026-04-10 19:46', 'April 2026'),
    ('126c5a63-cbe4-4905-913c-713369ebd5ca', 'VM/serverless infected with a high/critical severity threat', 'CRITICAL', 'IGNORED',  '2025-12-10', '2026-04-15 21:10', 'May 2026'),
    ('7077aee9-af8a-4d01-8b6e-464c3834309d', 'VM/serverless infected with a high/critical severity threat', 'CRITICAL', 'RESOLVED', '2025-12-10', '2026-04-23 15:45', 'May 2026'),
]

status_colors = {'RESOLVED': '375623', 'IGNORED': 'BF8F00'}

for ri, (issue_id, title, sev, status, created, resolved, month) in enumerate(issues, list_hdr_row + 1):
    row_fill = ZEBRA if ri % 2 == 0 else None
    s_fill   = PatternFill('solid', fgColor=status_colors.get(status, '666666'))
    for ci, val in enumerate([issue_id, title, sev, status, created, resolved, month], 1):
        ff = s_fill if ci == 4 else row_fill
        if ci == 3:
            fn = Font(name='Calibri', bold=True, size=10, color='C00000')
        elif ci == 4:
            fn = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        else:
            fn = Font(name='Calibri', size=10)
        h(ws, ri, ci, val, fill=ff, font=fn,
          align=LEFT_PAD if ci in (1, 2) else CTR, border=cell_border())
    ws.row_dimensions[ri].height = 20

# ── SECTION 8: Footer ─────────────────────────────────────────────────────────
footer_row = list_hdr_row + 1 + len(issues)
ws.merge_cells(f'A{footer_row}:G{footer_row}')
h(ws, footer_row, 1,
  f'Total: {len(issues)}  |  Resolved: 3  |  Ignored: 1  |  4 issues still IN_PROGRESS (not listed)',
  fill=HDR_DARK,
  font=Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
  align=LEFT_PAD)
ws.row_dimensions[footer_row].height = 20

# ── Column widths ──────────────────────────────────────────────────────────────
col_widths = [38, 40, 12, 12, 14, 24, 14]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'A5'

wb.save(r'C:\Users\PrashantKumar\csv2excel\CareFirst_Governance_Report.xlsx')
print('Done — 1 tab created: CRITICAL Issues')

