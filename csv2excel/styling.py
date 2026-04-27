"""Styling module — colours, fonts, cell formats, and print setup."""

from __future__ import annotations

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── Default brand palette ────────────────────────────────────────────────────

DEFAULT_PRIMARY = "1F4E79"     # dark blue
DEFAULT_SECONDARY = "2E75B6"   # medium blue
DEFAULT_ACCENT = "BDD7EE"      # light blue
DEFAULT_LIGHT_ROW = "F2F7FB"   # very light blue tint
DEFAULT_DARK_TEXT = "FFFFFF"    # white header text
DEFAULT_NEGATIVE = "FF4444"    # red for negative values
DEFAULT_POSITIVE = "27AE60"    # green for top values
DEFAULT_DUPE = "FFF3CD"        # yellow for duplicate rows


class BrandPalette:
    """Holds hex colour codes used throughout the workbook."""

    def __init__(self, primary: str = DEFAULT_PRIMARY):
        self.primary = primary.lstrip("#")
        self.secondary = DEFAULT_SECONDARY
        self.accent = DEFAULT_ACCENT
        self.light_row = DEFAULT_LIGHT_ROW
        self.header_font_color = DEFAULT_DARK_TEXT
        self.negative = DEFAULT_NEGATIVE
        self.positive = DEFAULT_POSITIVE
        self.dupe_highlight = DEFAULT_DUPE


# ── Named styles ─────────────────────────────────────────────────────────────

def register_styles(wb, palette: BrandPalette) -> None:
    """Create and register reusable NamedStyles on the workbook."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    # Header
    hdr = NamedStyle(name="header_style")
    hdr.font = Font(name="Calibri", bold=True, size=11, color=palette.header_font_color)
    hdr.fill = PatternFill("solid", fgColor=palette.primary)
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr.border = Border(bottom=Side(style="medium", color=palette.primary))
    wb.add_named_style(hdr)

    # Normal even row
    even = NamedStyle(name="even_row")
    even.font = Font(name="Calibri", size=10)
    even.border = border
    wb.add_named_style(even)

    # Odd (zebra) row
    odd = NamedStyle(name="odd_row")
    odd.font = Font(name="Calibri", size=10)
    odd.fill = PatternFill("solid", fgColor=palette.light_row)
    odd.border = border
    wb.add_named_style(odd)

    # Title
    title = NamedStyle(name="title_style")
    title.font = Font(name="Calibri", bold=True, size=22, color=palette.primary)
    title.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(title)

    # Subtitle
    sub = NamedStyle(name="subtitle_style")
    sub.font = Font(name="Calibri", size=13, italic=True, color="666666")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(sub)

    # KPI header
    kpi_h = NamedStyle(name="kpi_header")
    kpi_h.font = Font(name="Calibri", bold=True, size=11, color=palette.header_font_color)
    kpi_h.fill = PatternFill("solid", fgColor=palette.secondary)
    kpi_h.alignment = Alignment(horizontal="center", vertical="center")
    kpi_h.border = border
    wb.add_named_style(kpi_h)

    # KPI value
    kpi_v = NamedStyle(name="kpi_value")
    kpi_v.font = Font(name="Calibri", bold=True, size=14, color=palette.primary)
    kpi_v.alignment = Alignment(horizontal="center", vertical="center")
    kpi_v.border = border
    wb.add_named_style(kpi_v)


# ── Number format strings ────────────────────────────────────────────────────

FORMAT_MAP: dict[str, str] = {
    "currency": '$#,##0.00',
    "percentage": '0.00%',
    "integer": '#,##0',
    "float": '#,##0.00',
    "date": 'DD-MMM-YYYY',
    "text": "@",
}


def apply_number_format(cell, col_type: str) -> None:
    """Set the number format on a cell based on column type."""
    fmt = FORMAT_MAP.get(col_type)
    if fmt:
        cell.number_format = fmt


# ── Sheet-level helpers ──────────────────────────────────────────────────────

def auto_fit_columns(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """Adjust column widths based on content length."""
    for col_cells in ws.columns:
        lengths: list[int] = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if not lengths:
            continue
        best = max(lengths) + 2
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(best, max_width))


def setup_print(ws: Worksheet) -> None:
    """Configure print settings: landscape, fit-to-page, repeating headers."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # as many pages as needed vertically
    ws.print_title_rows = "1:1"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.right.text = "&D &T"


def apply_header_style(ws: Worksheet, col_count: int) -> None:
    """Style the first row as the header row and freeze it."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.style = "header_style"
    ws.freeze_panes = "A2"


def apply_zebra_stripes(ws: Worksheet, data_start_row: int, data_end_row: int) -> None:
    """Apply alternating row styles."""
    for row_idx in range(data_start_row, data_end_row + 1):
        style_name = "odd_row" if (row_idx - data_start_row) % 2 == 0 else "even_row"
        for cell in ws[row_idx]:
            cell.style = style_name


def highlight_negatives(ws: Worksheet, col_idx: int, data_start: int, data_end: int, palette: BrandPalette) -> None:
    """Color negative values red."""
    red_font = Font(name="Calibri", size=10, color=palette.negative, bold=True)
    for row_idx in range(data_start, data_end + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = red_font


def highlight_duplicates(ws: Worksheet, dupe_rows: set[int], data_start: int, col_count: int, palette: BrandPalette) -> None:
    """Highlight duplicate rows with a yellow fill."""
    fill = PatternFill("solid", fgColor=palette.dupe_highlight)
    for dupe_idx in dupe_rows:
        row_num = data_start + dupe_idx
        for col_idx in range(1, col_count + 1):
            ws.cell(row=row_num, column=col_idx).fill = fill
