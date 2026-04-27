#!/usr/bin/env python3
"""csv2excel — Convert raw CSV data into a polished, client-facing Excel workbook.

Usage:
    python -m csv2excel --input data.csv --output report.xlsx
    python -m csv2excel --input data.csv --title "Q4 Sales Report" --brand-color "#1F4E79"
    cat data.csv | python -m csv2excel --output report.xlsx      (stdin)
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from csv2excel.charts import add_pie_from_pivot, generate_charts
from csv2excel.data_processor import (
    analyze_columns,
    coerce_value,
    compute_numeric_stats,
    find_duplicates,
    read_csv,
)
from csv2excel.styling import (
    BrandPalette,
    apply_header_style,
    apply_number_format,
    apply_zebra_stripes,
    auto_fit_columns,
    highlight_duplicates,
    highlight_negatives,
    register_styles,
    setup_print,
)

logger = logging.getLogger("csv2excel")


# ── Cover Sheet ──────────────────────────────────────────────────────────────

def _build_cover(
    wb: Workbook,
    title: str,
    subtitle: str,
    prepared_by: str,
    logo_path: str | None,
    palette: BrandPalette,
) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = palette.primary

    # Logo
    current_row = 2
    if logo_path and Path(logo_path).is_file():
        img = XlImage(logo_path)
        img.width = 200
        img.height = 60
        ws.add_image(img, "B2")
        current_row = 6
    else:
        current_row = 4

    # Title
    cell = ws.cell(row=current_row, column=2, value=title)
    cell.style = "title_style"
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row, end_column=6)

    # Subtitle
    current_row += 2
    cell = ws.cell(row=current_row, column=2, value=subtitle)
    cell.style = "subtitle_style"
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row, end_column=6)

    # Metadata
    current_row += 3
    meta = [
        ("Generated:", datetime.now().strftime("%d-%b-%Y %H:%M:%S")),
        ("Prepared by:", prepared_by),
    ]
    for label, value in meta:
        ws.cell(row=current_row, column=2, value=label).font = \
            ws.cell(row=current_row, column=2).font.copy(bold=True)
        ws.cell(row=current_row, column=3, value=value)
        current_row += 1

    # Disclaimer
    current_row += 2
    disc = ws.cell(
        row=current_row, column=2,
        value="CONFIDENTIAL — This report is intended for authorized recipients only.",
    )
    disc.font = disc.font.copy(italic=True, color="999999", size=9)
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row, end_column=6)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    for c in "CDEF":
        ws.column_dimensions[c].width = 18
    setup_print(ws)


# ── Data Sheet ───────────────────────────────────────────────────────────────

def _build_data_sheet(
    wb: Workbook,
    headers: list[str],
    rows: list[list[str]],
    col_meta: list[dict[str, Any]],
    dupe_rows: set[int],
    palette: BrandPalette,
) -> None:
    ws = wb.create_sheet("Data")
    ws.sheet_properties.tabColor = palette.secondary
    col_count = len(headers)
    data_start = 2
    data_end = data_start + len(rows) - 1

    # Write headers
    for ci, header in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=header)

    # Write data rows
    for ri, row in enumerate(rows):
        for ci, meta in enumerate(col_meta):
            raw = row[ci] if ci < len(row) else ""
            value = coerce_value(raw, meta["type"])
            cell = ws.cell(row=data_start + ri, column=ci + 1, value=value)
            apply_number_format(cell, meta["type"])

    # Styling
    apply_header_style(ws, col_count)
    if len(rows) <= 10000:  # skip zebra on huge files for performance
        apply_zebra_stripes(ws, data_start, data_end)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{data_end}"

    # Highlight negatives
    for meta in col_meta:
        if meta["type"] in ("currency", "integer", "float"):
            highlight_negatives(ws, meta["index"] + 1, data_start, data_end, palette)

    # Highlight duplicates
    if dupe_rows:
        highlight_duplicates(ws, dupe_rows, data_start, col_count, palette)

    # Data validation for categorical columns
    for meta in col_meta:
        if meta.get("is_categorical"):
            samples = sorted(set(
                row[meta["index"]] for row in rows
                if meta["index"] < len(row) and row[meta["index"]].strip()
            ))
            if 2 <= len(samples) <= 20:
                formula = '"' + ",".join(samples) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv.error = "Please select from the dropdown list"
                dv.errorTitle = "Invalid Entry"
                col_letter = get_column_letter(meta["index"] + 1)
                dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
                ws.add_data_validation(dv)

    # Named range
    from openpyxl.workbook.defined_name import DefinedName
    range_str = f"Data!$A$1:${get_column_letter(col_count)}${data_end}"
    dn = DefinedName("DataRange", attr_text=range_str)
    wb.defined_names.add(dn)

    auto_fit_columns(ws)
    setup_print(ws)


# ── Summary / Dashboard Sheet ───────────────────────────────────────────────

def _build_summary(
    wb: Workbook,
    headers: list[str],
    rows: list[list[str]],
    col_meta: list[dict[str, Any]],
    palette: BrandPalette,
) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "27AE60"
    data_ws = wb["Data"]

    row_cursor = 2

    # ── KPI Cards ────────────────────────────────────────────────────────
    ws.cell(row=row_cursor, column=2, value="KEY METRICS").style = "kpi_header"
    ws.merge_cells(start_row=row_cursor, start_column=2,
                   end_row=row_cursor, end_column=7)
    row_cursor += 1

    ws.cell(row=row_cursor, column=2, value="Total Rows").style = "kpi_header"
    ws.cell(row=row_cursor, column=3, value=len(rows)).style = "kpi_value"

    ws.cell(row=row_cursor, column=4, value="Total Columns").style = "kpi_header"
    ws.cell(row=row_cursor, column=5, value=len(headers)).style = "kpi_value"
    row_cursor += 2

    # Stats per numeric column
    numeric_meta = [m for m in col_meta if m["type"] in ("currency", "integer", "float", "percentage")]
    if numeric_meta:
        stat_headers = ["Column", "Sum", "Average", "Min", "Max", "Count"]
        for ci, sh in enumerate(stat_headers, 2):
            ws.cell(row=row_cursor, column=ci, value=sh).style = "kpi_header"
        row_cursor += 1

        for meta in numeric_meta:
            values = [
                coerce_value(r[meta["index"]], meta["type"])
                for r in rows if meta["index"] < len(r)
            ]
            stats = compute_numeric_stats(values)
            ws.cell(row=row_cursor, column=2, value=meta["header"])
            ws.cell(row=row_cursor, column=3, value=stats["sum"])
            ws.cell(row=row_cursor, column=4, value=stats["avg"])
            ws.cell(row=row_cursor, column=5, value=stats["min"])
            ws.cell(row=row_cursor, column=6, value=stats["max"])
            ws.cell(row=row_cursor, column=7, value=stats["count"])
            # format currency columns
            if meta["type"] == "currency":
                for c in (3, 4, 5, 6):
                    apply_number_format(ws.cell(row=row_cursor, column=c), "currency")
            elif meta["type"] == "percentage":
                for c in (3, 4, 5, 6):
                    apply_number_format(ws.cell(row=row_cursor, column=c), "percentage")
            row_cursor += 1
        row_cursor += 1

    # ── Pivot tables for categorical columns ─────────────────────────────
    cat_meta = [m for m in col_meta if m.get("is_categorical")]
    for meta in cat_meta[:4]:  # limit to 4 pivots
        ws.cell(row=row_cursor, column=2,
                value=f"Breakdown: {meta['header']}").style = "kpi_header"
        ws.merge_cells(start_row=row_cursor, start_column=2,
                       end_row=row_cursor, end_column=4)
        row_cursor += 1

        pivot_start = row_cursor
        ws.cell(row=row_cursor, column=2, value=meta["header"]).style = "kpi_header"
        ws.cell(row=row_cursor, column=3, value="Count").style = "kpi_header"
        row_cursor += 1

        counts = Counter(
            r[meta["index"]] for r in rows
            if meta["index"] < len(r) and r[meta["index"]].strip()
        )
        for val, cnt in counts.most_common():
            ws.cell(row=row_cursor, column=2, value=val)
            ws.cell(row=row_cursor, column=3, value=cnt)
            row_cursor += 1

        pivot_end = row_cursor - 1
        # Pie chart from this pivot
        anchor = f"E{pivot_start}"
        add_pie_from_pivot(ws, meta["header"], 2, 3, pivot_start, pivot_end, anchor)
        row_cursor += 2

    # ── Auto Charts ──────────────────────────────────────────────────────
    row_cursor += 1
    data_start = 2
    data_end = data_start + len(rows) - 1
    row_cursor = generate_charts(ws, data_ws, col_meta, data_start, data_end, row_cursor)

    auto_fit_columns(ws)
    setup_print(ws)


# ── Data Dictionary Sheet ────────────────────────────────────────────────────

def _build_data_dictionary(
    wb: Workbook,
    col_meta: list[dict[str, Any]],
    palette: BrandPalette,
) -> None:
    ws = wb.create_sheet("Data Dictionary")
    ws.sheet_properties.tabColor = "E67E22"

    dict_headers = [
        "Column Name", "Detected Type", "Total Count", "Null Count",
        "Unique Count", "Sample Values",
    ]
    for ci, h in enumerate(dict_headers, 1):
        ws.cell(row=1, column=ci, value=h)

    for ri, meta in enumerate(col_meta, 2):
        ws.cell(row=ri, column=1, value=meta["header"])
        ws.cell(row=ri, column=2, value=meta["type"])
        ws.cell(row=ri, column=3, value=meta["total_count"])
        ws.cell(row=ri, column=4, value=meta["null_count"])
        ws.cell(row=ri, column=5, value=meta["unique_count"])
        ws.cell(row=ri, column=6, value=", ".join(str(v) for v in meta["sample_values"]))

    apply_header_style(ws, len(dict_headers))
    apply_zebra_stripes(ws, 2, len(col_meta) + 1)
    auto_fit_columns(ws)
    setup_print(ws)


# ── Main entry point ─────────────────────────────────────────────────────────

def build_workbook(
    source: str | Path,
    output: str | Path,
    title: str = "Data Report",
    subtitle: str = "",
    prepared_by: str = "Automated System",
    brand_color: str = "1F4E79",
    logo_path: str | None = None,
) -> Path:
    """End-to-end: read CSV → build Excel workbook → save."""
    palette = BrandPalette(primary=brand_color)
    headers, rows = read_csv(source)
    col_meta = analyze_columns(headers, rows)
    dupe_rows = find_duplicates(rows)

    logger.info("Detected types: %s",
                {m['header']: m['type'] for m in col_meta})
    if dupe_rows:
        logger.info("Found %d duplicate row(s)", len(dupe_rows))

    wb = Workbook()
    register_styles(wb, palette)

    _build_cover(wb, title, subtitle or f"Generated from {Path(source).name if isinstance(source, (str, Path)) and Path(source).is_file() else 'CSV input'}",
                 prepared_by, logo_path, palette)
    _build_data_sheet(wb, headers, rows, col_meta, dupe_rows, palette)
    _build_summary(wb, headers, rows, col_meta, palette)
    _build_data_dictionary(wb, col_meta, palette)

    out = Path(output)
    wb.save(out)
    logger.info("Workbook saved → %s", out.resolve())
    return out


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="csv2excel",
        description="Transform raw CSV data into a polished, client-facing Excel workbook.",
    )
    parser.add_argument("--input", "-i", required=False,
                        help="Path to the input CSV file. Omit to read from stdin.")
    parser.add_argument("--output", "-o", default="report.xlsx",
                        help="Output Excel file path (default: report.xlsx)")
    parser.add_argument("--title", "-t", default="Data Report",
                        help="Report title on the cover sheet")
    parser.add_argument("--subtitle", default="",
                        help="Report subtitle")
    parser.add_argument("--prepared-by", default="Automated System",
                        help="Name shown on the cover sheet")
    parser.add_argument("--brand-color", default="1F4E79",
                        help="Primary brand hex color (default: 1F4E79)")
    parser.add_argument("--logo", default=None,
                        help="Path to a logo image (png/jpg)")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    if args.input:
        source: str | Path = Path(args.input)
        if not source.is_file():
            logger.error("File not found: %s", source)
            sys.exit(1)
    else:
        logger.info("Reading CSV from stdin…")
        source = sys.stdin.read()
        if not source.strip():
            logger.error("No data received on stdin")
            sys.exit(1)

    out = build_workbook(
        source=source,
        output=args.output,
        title=args.title,
        subtitle=args.subtitle,
        prepared_by=args.prepared_by,
        brand_color=args.brand_color.lstrip("#"),
        logo_path=args.logo,
    )
    print(f"\n✅ Report saved: {out.resolve()}")


if __name__ == "__main__":
    main()
