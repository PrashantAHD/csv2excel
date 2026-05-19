#!/usr/bin/env python3
"""wiz_report — Generate a month-on-month governance Excel report from a Wiz CSV export.

Usage (non-interactive):
    python -m csv2excel.wiz_report -i issues.csv -o report.xlsx --mode new
    python -m csv2excel.wiz_report -i issues.csv -o report.xlsx --mode append
    python -m csv2excel.wiz_report -i issues.csv -o report.xlsx --snapshots "Mar 26,Apr 26,May 26"

Usage (interactive):
    python -m csv2excel.wiz_report
    wiz-report                          # after pip install
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%Y-%m-%dT%H:%M:%SZ",
    "%Y-%m-%d %H:%M:%S.%f +0000 UTC",
    "%Y-%m-%d %H:%M:%S +0000 UTC",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
]

REQUIRED_COLUMNS = {"Issue ID", "Title", "Severity", "Status", "Created At", "Resolved Time"}

SEVERITY_COLORS = {
    "CRITICAL": "C00000",
    "HIGH":     "C55A11",
    "MEDIUM":   "7030A0",
    "LOW":      "2E75B6",
    "INFO":     "595959",
}
SEVERITY_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]

# ── Shared style constants ─────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color="BDD7EE")
    return Border(top=s, bottom=s, left=s, right=s)


HDR_DARK  = PatternFill("solid", fgColor="1F4E79")
HDR_GREEN = PatternFill("solid", fgColor="375623")
ZEBRA     = PatternFill("solid", fgColor="EBF3FB")
CTR_WRAP  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTR       = Alignment(horizontal="center", vertical="center")
LEFT_PAD  = Alignment(horizontal="left",   vertical="center", indent=1)
WHITE_BOLD = Font(name="Calibri", bold=True,  size=11, color="FFFFFF")
DARK_BOLD  = Font(name="Calibri", bold=True,  size=11, color="1F4E79")
REG        = Font(name="Calibri", size=11)


def _wc(ws, row: int, col: int, val=None, fill=None, font=None, align=None, border=None):
    """Write a cell with optional formatting."""
    c = ws.cell(row=row, column=col, value=val)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c


# ── Date helpers ───────────────────────────────────────────────────────────────

def _parse_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    s = s.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def default_snapshots() -> list[tuple[str, datetime]]:
    """Return (label, datetime) for the 15th of each of the last 3 months."""
    today = datetime.today()
    snaps: list[tuple[str, datetime]] = []
    for offset in range(2, -1, -1):
        month = today.month - offset
        year  = today.year
        while month <= 0:
            month += 12
            year  -= 1
        dt = datetime(year, month, 15, 23, 59, 59)
        snaps.append((dt.strftime("%b 15, %Y"), dt))  # always the 15th
    return snaps


def parse_snapshots(text: str) -> list[tuple[str, datetime]]:
    """Parse 'Mar 26,Apr 26,May 26' into [(label, datetime), ...].
    Label is always normalised to 'Mon 15, YYYY' so the 15th is explicit.
    """
    result: list[tuple[str, datetime]] = []
    for token in text.split(","):
        token = token.strip()
        try:
            dt = datetime.strptime(token, "%b %y")
            dt = datetime(dt.year, dt.month, 15, 23, 59, 59)
            result.append((dt.strftime("%b 15, %Y"), dt))  # always the 15th
        except ValueError:
            print(f"  Warning: could not parse snapshot date '{token}' — skipped.")
    return result


# ── Metrics ────────────────────────────────────────────────────────────────────

def compute_metrics(
    rows: list[dict],
    snapshots: list[tuple[str, datetime]],
) -> dict:
    """Compute open / resolved / ignored counts and resolution % at each snapshot."""
    out: dict = {
        "labels":   [s[0] for s in snapshots],
        "total":    [],
        "open":     [],
        "resolved": [],
        "ignored":  [],
        "pct":      [],
    }
    for _, sd in snapshots:
        total    = sum(1 for r in rows
                       if (c := _parse_dt(r.get("Created At"))) and c <= sd)
        open_    = sum(1 for r in rows
                       if (c := _parse_dt(r.get("Created At"))) and c <= sd
                       and ((rv := _parse_dt(r.get("Resolved Time"))) is None or rv > sd))
        resolved = sum(1 for r in rows
                       if r.get("Status", "").strip() == "RESOLVED"
                       and (rv := _parse_dt(r.get("Resolved Time"))) and rv <= sd)
        ignored  = sum(1 for r in rows
                       if r.get("Status", "").strip() == "REJECTED"
                       and (rv := _parse_dt(r.get("Resolved Time"))) and rv <= sd)
        ri = resolved + ignored
        out["total"].append(total)
        out["open"].append(open_)
        out["resolved"].append(resolved)
        out["ignored"].append(ignored)
        out["pct"].append(f"{ri / total * 100:.1f}%" if total else "0%")
    return out


def _diff(a, b) -> str:
    """Format the signed change between two values (numeric or pct string)."""
    try:
        fa = float(str(a).rstrip("%"))
        fb = float(str(b).rstrip("%"))
        d  = fb - fa
        is_pct = isinstance(a, str) and "%" in a
        s = "+" if d >= 0 else ""
        return f"{s}{d:.1f}%" if is_pct else f"{s}{int(d)}"
    except (ValueError, TypeError):
        return "-"


def read_metrics_from_existing_tab(
    wb: Workbook,
    severity: str,
    n_snapshots: int = 3,
) -> dict | None:
    """Read back metrics from an already-built severity tab in the workbook.
    Used in append mode so severities not in the new CSV still appear in Summary.
    """
    tab_name = f"{severity} Issues"
    if tab_name not in wb.sheetnames:
        return None
    ws = wb[tab_name]

    # Read snapshot labels from row 4, cols 2..(1+n_snapshots)
    labels = []
    for ci in range(2, 2 + n_snapshots):
        raw = str(ws.cell(row=4, column=ci).value or "")
        label = raw.replace("\n(Baseline)", "").replace("\n(Latest)", "").strip()
        labels.append(label)

    # Metric rows 5-9: Total | Open | Resolved | Ignored | Resolution%
    keys = ["total", "open", "resolved", "ignored", "pct"]
    result: dict = {
        "labels":   labels,
        "total":    [], "open": [], "resolved": [], "ignored": [], "pct": [],
    }
    for offset, key in enumerate(keys):
        row_vals = []
        for ci in range(2, 2 + n_snapshots):
            val = ws.cell(row=5 + offset, column=ci).value
            row_vals.append(val if val is not None else ("0%" if key == "pct" else 0))
        result[key] = row_vals
    return result


# ── Severity tab ───────────────────────────────────────────────────────────────

def build_severity_tab(
    wb: Workbook,
    severity: str,
    rows: list[dict],
    snapshots: list[tuple[str, datetime]],
) -> None:
    tab_name = f"{severity} Issues"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    color = SEVERITY_COLORS.get(severity, "595959")
    ws.sheet_properties.tabColor = color
    sev_fill = PatternFill("solid", fgColor=color)
    n = len(snapshots)
    labels = [s[0] for s in snapshots]

    # Column layout: Metric | snap×n | vs Baseline | Monthly×(n-1)
    # Capped at 7 columns to match existing report format
    ncols = min(1 + n + 1 + (n - 1), 7)

    # ── Title rows ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    _wc(ws, 1, 1, f"{severity} Severity — Resolution Progress (Month-on-Month)",
        fill=sev_fill, font=Font(name="Calibri", bold=True, size=14, color="FFFFFF"), align=CTR_WRAP)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    _wc(ws, 2, 1, f"Exec Summary  |  Severity: {severity}  |  As of {labels[-1]}",
        fill=PatternFill("solid", fgColor="3C3C3C"),
        font=Font(name="Calibri", italic=True, size=10, color="FFFFFF"), align=LEFT_PAD)
    ws.row_dimensions[2].height = 18

    # ── Section label row ──────────────────────────────────────────────────
    snap_end = 1 + n   # last snapshot column index
    prog_start = snap_end + 1
    ws.merge_cells(f"A3:{get_column_letter(snap_end)}3")
    _wc(ws, 3, 1, "SNAPSHOT DATA",
        fill=PatternFill("solid", fgColor="2E75B6"), font=WHITE_BOLD, align=CTR)
    for ci in range(2, snap_end + 1):
        ws.cell(row=3, column=ci).fill = PatternFill("solid", fgColor="2E75B6")
    ws.merge_cells(f"{get_column_letter(prog_start)}3:{get_column_letter(ncols)}3")
    _wc(ws, 3, prog_start, "PROGRESS TRACKING", fill=HDR_GREEN, font=WHITE_BOLD, align=CTR)
    for ci in range(prog_start + 1, ncols + 1):
        ws.cell(row=3, column=ci).fill = HDR_GREEN
    ws.row_dimensions[3].height = 22

    # ── Column headers ─────────────────────────────────────────────────────
    snap_labels = (
        [f"{labels[0]}\n(Baseline)"]
        + labels[1:-1]
        + ([f"{labels[-1]}\n(Latest)"] if n > 1 else [])
    )
    progress_headers = [f"vs Baseline\n({labels[0]} → {labels[-1]})"]
    for i in range(n - 1):
        progress_headers.append(f"Monthly\n{labels[i]} → {labels[i+1]}")

    all_headers = ["Metric"] + snap_labels + progress_headers
    all_headers = all_headers[:ncols]

    for ci, ch in enumerate(all_headers, 1):
        fill = HDR_DARK if ci <= snap_end else HDR_GREEN
        _wc(ws, 4, ci, ch, fill=fill, font=WHITE_BOLD, align=CTR_WRAP, border=_thin_border())
    ws.row_dimensions[4].height = 40

    # ── Metrics rows ───────────────────────────────────────────────────────
    m = compute_metrics(rows, snapshots)
    metric_defs = [
        ("Total Issues Created",                  m["total"],    PatternFill("solid", fgColor="F2F2F2")),
        ("Open / In Progress",                    m["open"],     PatternFill("solid", fgColor="FCE4D6")),
        ("Resolved  (Cumulative)",                m["resolved"], None),
        ("Ignored / Accepted Risk  (Cumulative)", m["ignored"],  None),
        ("Resolution %  (Resolved + Ignored)",    m["pct"],      PatternFill("solid", fgColor="FFF2CC")),
    ]

    for ri, (label, vals, row_fill) in enumerate(metric_defs, 5):
        is_pct = "Resolution %" in label
        vs_base = _diff(vals[0], vals[-1])
        monthly = [_diff(vals[i], vals[i + 1]) for i in range(n - 1)]
        all_vals = ([label] + list(vals) + [vs_base] + monthly)[:ncols]
        for ci, val in enumerate(all_vals, 1):
            ff = row_fill or (ZEBRA if ri % 2 == 0 else None)
            fn = Font(name="Calibri", bold=True, size=11, color="7F6000") if is_pct \
                 else (DARK_BOLD if ci == 1 else REG)
            _wc(ws, ri, ci, val, fill=ff, font=fn,
                align=LEFT_PAD if ci == 1 else CTR, border=_thin_border())
        ws.row_dimensions[ri].height = 20

    # ── Footnote ───────────────────────────────────────────────────────────
    fn_row = 5 + len(metric_defs)
    ws.merge_cells(f"A{fn_row}:{get_column_letter(ncols)}{fn_row}")
    _wc(ws, fn_row, 1, "* Snapshots taken as of End-of-Day (23:59:59) on each date.",
        font=Font(name="Calibri", italic=True, size=9, color="808080"), align=LEFT_PAD)
    ws.row_dimensions[fn_row].height = 16

    # ── Resolved / ignored issue detail list ──────────────────────────────
    ri_issues = sorted(
        [r for r in rows if r.get("Status", "").strip() in ("RESOLVED", "REJECTED")],
        key=lambda r: _parse_dt(r.get("Resolved Time")) or datetime.max,
    )

    def snap_month(dt: datetime | None) -> str:
        if dt is None:
            return "-"
        for lbl, sd in snapshots:
            if dt <= sd:
                try:
                    return datetime.strptime(lbl, "%b %d, %Y").strftime("%B %Y")
                except ValueError:
                    return lbl
        return dt.strftime("%B %Y")

    section_row = fn_row + 2
    ws.merge_cells(f"A{section_row}:{get_column_letter(ncols)}{section_row}")
    _wc(ws, section_row, 1, "Resolved Issues Detail  (Resolved + Ignored)",
        fill=PatternFill("solid", fgColor="3C3C3C"),
        font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"), align=LEFT_PAD)
    ws.row_dimensions[section_row].height = 22

    ws.merge_cells(f"A{section_row+1}:{get_column_letter(ncols)}{section_row+1}")
    _wc(ws, section_row + 1, 1,
        "IGNORED = team has accepted the risk and committed to a fix-by date",
        fill=PatternFill("solid", fgColor="595959"),
        font=Font(name="Calibri", italic=True, size=9, color="FFFFFF"), align=LEFT_PAD)
    ws.row_dimensions[section_row + 1].height = 16

    list_hdr_row = section_row + 2
    list_headers = ["Issue ID", "Title", "Severity", "Status", "Created At", "Resolved Time", "Snapshot Month"]
    for ci, ch in enumerate(list_headers, 1):
        _wc(ws, list_hdr_row, ci, ch,
            fill=HDR_DARK, font=WHITE_BOLD, align=CTR_WRAP, border=_thin_border())
    ws.row_dimensions[list_hdr_row].height = 28

    status_colors = {"RESOLVED": "375623", "REJECTED": "BF8F00"}
    for ri, r in enumerate(ri_issues, list_hdr_row + 1):
        zfill = ZEBRA if ri % 2 == 0 else None
        status = r.get("Status", "").strip()
        display_status = "IGNORED" if status == "REJECTED" else status
        s_fill = PatternFill("solid", fgColor=status_colors.get(status, "666666"))
        created_dt  = _parse_dt(r.get("Created At"))
        resolved_dt = _parse_dt(r.get("Resolved Time"))
        vals = [
            r.get("Issue ID", "").strip(),
            r.get("Title", "").strip(),
            severity,
            display_status,
            created_dt.strftime("%Y-%m-%d") if created_dt else "",
            resolved_dt.strftime("%Y-%m-%d %H:%M") if resolved_dt else "",
            snap_month(resolved_dt),
        ]
        for ci, val in enumerate(vals, 1):
            ff = s_fill if ci == 4 else zfill
            if ci == 3:
                fn = Font(name="Calibri", bold=True, size=10, color=color)
            elif ci == 4:
                fn = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            else:
                fn = Font(name="Calibri", size=10)
            _wc(ws, ri, ci, val, fill=ff, font=fn,
                align=LEFT_PAD if ci in (1, 2) else CTR, border=_thin_border())
        ws.row_dimensions[ri].height = 18

    footer_row = list_hdr_row + 1 + len(ri_issues)
    res_all = sum(1 for r in rows if r.get("Status", "").strip() == "RESOLVED")
    ign_all = sum(1 for r in rows if r.get("Status", "").strip() == "REJECTED")
    ws.merge_cells(f"A{footer_row}:{get_column_letter(len(list_headers))}{footer_row}")
    _wc(ws, footer_row, 1,
        f"Total: {len(ri_issues)}  |  Resolved: {res_all}  |  Ignored: {ign_all}"
        f"  |  {len(rows) - len(ri_issues)} issue(s) still Open / In-Progress (not listed)",
        fill=HDR_DARK, font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"), align=LEFT_PAD)
    ws.row_dimensions[footer_row].height = 20

    for ci, w in enumerate([38, 42, 10, 12, 14, 22, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A5"


# ── Summary tab ────────────────────────────────────────────────────────────────

def build_summary_tab(
    wb: Workbook,
    severity_metrics: dict[str, dict],
    snapshots: list[tuple[str, datetime]],
    title: str = "Cloud Security — Governance Report",
) -> None:
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "1F4E79"

    labels = [s[0] for s in snapshots]
    latest    = labels[-1]
    baseline  = labels[0]

    ws.merge_cells("A1:H1")
    _wc(ws, 1, 1, title,
        fill=HDR_DARK, font=Font(name="Calibri", bold=True, size=16, color="FFFFFF"), align=CTR_WRAP)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    _wc(ws, 2, 1, f"Exec Summary  |  Snapshot: {latest}  |  All Severities",
        fill=PatternFill("solid", fgColor="3C3C3C"),
        font=Font(name="Calibri", italic=True, size=10, color="FFFFFF"), align=LEFT_PAD)
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:E3")
    _wc(ws, 3, 1, f"LATEST SNAPSHOT  ({latest})",
        fill=PatternFill("solid", fgColor="2E75B6"), font=WHITE_BOLD, align=CTR)
    for ci in range(2, 6):
        ws.cell(row=3, column=ci).fill = PatternFill("solid", fgColor="2E75B6")
    ws.merge_cells("F3:H3")
    _wc(ws, 3, 6, f"PROGRESS vs BASELINE  ({baseline})",
        fill=HDR_GREEN, font=WHITE_BOLD, align=CTR)
    for ci in range(7, 9):
        ws.cell(row=3, column=ci).fill = HDR_GREEN
    ws.row_dimensions[3].height = 22

    col_headers = [
        "Severity", "Total\nIssues", "Open /\nIn Progress",
        "Resolved\n(Cumulative)", "Ignored /\nAccepted Risk",
        "Resolution %\n(Resolved + Ignored)",
        f"vs Baseline\n({baseline} → {latest})",
        "Trend",
    ]
    for ci, ch in enumerate(col_headers, 1):
        fill = HDR_DARK if ci <= 5 else HDR_GREEN
        _wc(ws, 4, ci, ch, fill=fill, font=WHITE_BOLD, align=CTR_WRAP, border=_thin_border())
    ws.row_dimensions[4].height = 44

    row = 5
    for sev in SEVERITY_ORDER:
        if sev not in severity_metrics:
            continue
        m     = severity_metrics[sev]
        color = SEVERITY_COLORS.get(sev, "595959")
        sev_fill = PatternFill("solid", fgColor=color)
        zfill    = ZEBRA if row % 2 == 0 else None

        p0 = float(str(m["pct"][0]).rstrip("%"))
        p1 = float(str(m["pct"][-1]).rstrip("%"))
        vs_base = f"{'+' if p1 - p0 >= 0 else ''}{p1 - p0:.1f}%"

        vals = [
            sev,
            m["total"][-1],
            m["open"][-1],
            m["resolved"][-1],
            m["ignored"][-1],
            m["pct"][-1],
            vs_base,
            "↓ Open issues declining",
        ]
        for ci, val in enumerate(vals, 1):
            if ci == 1:
                _wc(ws, row, ci, val, fill=sev_fill,
                    font=Font(name="Calibri", bold=True, size=12, color="FFFFFF"),
                    align=CTR, border=_thin_border())
            elif ci == 6:
                _wc(ws, row, ci, val, fill=PatternFill("solid", fgColor="FFF2CC"),
                    font=Font(name="Calibri", bold=True, size=12, color="7F6000"),
                    align=CTR, border=_thin_border())
            elif ci == 7:
                _wc(ws, row, ci, val, fill=HDR_GREEN,
                    font=Font(name="Calibri", bold=True, size=12, color="FFFFFF"),
                    align=CTR, border=_thin_border())
            elif ci == 3:
                _wc(ws, row, ci, val, fill=PatternFill("solid", fgColor="FCE4D6"),
                    font=Font(name="Calibri", bold=True, size=12, color="C00000"),
                    align=CTR, border=_thin_border())
            elif ci == 8:
                _wc(ws, row, ci, val, fill=PatternFill("solid", fgColor="E2EFDA"),
                    font=Font(name="Calibri", italic=True, size=10, color="375623"),
                    align=CTR, border=_thin_border())
            else:
                _wc(ws, row, ci, val, fill=zfill,
                    font=Font(name="Calibri", bold=True, size=12),
                    align=CTR, border=_thin_border())
        ws.row_dimensions[row].height = 28
        row += 1

    note_row = row + 1
    ws.merge_cells(f"A{note_row}:H{note_row}")
    _wc(ws, note_row, 1,
        "* Snapshots taken as of End-of-Day (23:59:59) on each date."
        "  |  See individual severity tabs for month-on-month breakdown.",
        font=Font(name="Calibri", italic=True, size=9, color="808080"), align=LEFT_PAD)
    ws.row_dimensions[note_row].height = 16

    for ci, w in enumerate([14, 10, 14, 14, 16, 24, 18, 26], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A5"


# ── Watermark ──────────────────────────────────────────────────────────────────

def apply_watermark(wb: Workbook, org: str = "PROPRIETARY") -> None:
    header = f'&"Calibri,Bold"&10&KA6A6A6CONFIDENTIAL \u2014 {org}'
    footer_l = '&"Calibri,Regular"&9&K808080Confidential \u2014 Proprietary'
    footer_r = '&"Calibri,Regular"&9&K808080Page &P of &N'
    for ws in wb.worksheets:
        ws.oddHeader.center.text = header
        ws.oddFooter.left.text   = footer_l
        ws.oddFooter.right.text  = footer_r


# ── CSV loading & validation ───────────────────────────────────────────────────

def load_wiz_csv(path: str) -> list[dict]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        print("Error: CSV file is empty.")
        sys.exit(1)
    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        print(f"Error: CSV is missing required columns: {', '.join(sorted(missing))}")
        print(f"  Found columns: {', '.join(rows[0].keys())}")
        sys.exit(1)
    return rows


def _parse_csv_text(text: str) -> list[dict]:
    """Parse raw CSV text (with header row) into validated row dicts."""
    reader = csv.DictReader(text.splitlines())
    rows = list(reader)
    if not rows:
        print("Error: CSV text is empty.")
        sys.exit(1)
    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        print(f"Error: CSV is missing required columns: {', '.join(sorted(missing))}")
        print(f"  Found columns: {', '.join(rows[0].keys())}")
        sys.exit(1)
    return rows


# ── Interactive prompts ────────────────────────────────────────────────────────

def _prompt_inputs(args: argparse.Namespace):
    """Collect any missing inputs interactively."""
    print("\n=== Wiz Governance Report Generator ===\n")

    # 1. Mode FIRST — so we know whether to ask for CSVs or just a workbook path
    if not args.mode:
        choice = input(
            "(N)ew report, (A)ppend to existing, or (R)emove a severity? [N]: "
        ).strip().upper() or "N"
        if choice.startswith("A"):
            args.mode = "append"
        elif choice.startswith("R"):
            args.mode = "remove"
        else:
            args.mode = "new"

    # ── Remove mode: only needs workbook path + severities to remove ──────────
    if args.mode == "remove":
        if not args.output or args.output == "report.xlsx":
            prompt = "Enter path to existing Excel workbook: "
            out = input(prompt).strip().strip('"')
            while not out or not Path(out).is_file():
                print(f"  File not found: {out or '(empty)'}")
                out = input(prompt).strip().strip('"')
            args.output = out
        if not args.remove:
            raw = input(
                "Severities to remove (e.g. MEDIUM  or  MEDIUM,LOW): "
            ).strip().upper()
            args.remove = [s.strip() for s in raw.split(",") if s.strip()]
        return args

    # ── New / Append mode ─────────────────────────────────────────────────────

    # 2. Input CSV(s) — file path OR inline paste, one per severity export
    collected_rows: list[dict] = []

    def _ask_csv(label: str) -> list[dict]:
        """Prompt for one CSV — accept a file path or pasted inline data."""
        print(f"\n{label}")
        while True:
            p = input("  File path (or type 'paste' to enter CSV data inline): ").strip().strip('"')
            if p.lower() == "paste":
                print("  Paste CSV rows below (include the header line). Enter a blank line to finish:")
                lines: list[str] = []
                while True:
                    line = input()
                    if not line:
                        break
                    lines.append(line)
                if not lines:
                    print("  Nothing entered — try again.")
                    continue
                return _parse_csv_text("\n".join(lines))
            elif Path(p).is_file():
                return load_wiz_csv(p)
            else:
                print(f"  File not found: '{p or '(empty)'}'. Enter a valid path or type 'paste'.")

    collected_rows.extend(_ask_csv("Wiz CSV export #1:"))
    while True:
        more = input("\nAdd another CSV (e.g. for a different severity)? (Y/N) [N]: "
                     ).strip().upper() or "N"
        if not more.startswith("Y"):
            break
        collected_rows.extend(_ask_csv(f"Wiz CSV export #{len(set(r.get('Severity','') for r in collected_rows)) + 1}:"))

    args._pre_loaded_rows = collected_rows  # type: ignore[attr-defined]
    args.input = ["<inline>"]  # sentinel so downstream checks pass

    # 3. Output path — prompt label reflects the chosen mode
    if not args.output or args.output == "report.xlsx":
        if args.mode == "append":
            prompt = "Enter path to existing Excel workbook: "
            out = input(prompt).strip().strip('"')
            while not out or not Path(out).is_file():
                print(f"  File not found: {out or '(empty)'}")
                out = input(prompt).strip().strip('"')
            args.output = out
        else:
            out = input("Enter output file name [report.xlsx]: ").strip().strip('"')
            args.output = out or "report.xlsx"

    # 4. Snapshot dates
    if not args.snapshots:
        default_snaps = default_snapshots()
        default_label = ", ".join(s[0] for s in default_snaps)
        raw = input(
            f"Snapshot months as 'Mon YY' separated by commas [{default_label}]: "
        ).strip()
        args.snapshots = raw if raw else None

    return args


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        prog="wiz-report",
        description="Generate a month-on-month governance Excel report from a Wiz CSV export.",
    )
    parser.add_argument("--input", "-i", nargs="+", default=None, metavar="CSV",
                        help="One or more Wiz CSV export files (e.g. one per severity).")
    parser.add_argument("--output",    "-o", default="report.xlsx",
                        help="Output Excel file path (default: report.xlsx).")
    parser.add_argument("--mode",      "-m", choices=["new", "append", "remove"], default=None,
                        help="'new' creates a fresh workbook; 'append' updates an existing one; "
                             "'remove' deletes severity tab(s) and rebuilds Summary. "
                             "Prompted if omitted.")
    parser.add_argument("--remove",    "-r", nargs="+", metavar="SEVERITY",
                        help="Severities to remove when --mode remove is used "
                             "(e.g. --remove MEDIUM  or  --remove MEDIUM LOW).")
    parser.add_argument("--snapshots", "-s", default=None,
                        help="Comma-separated snapshot months, e.g. 'Mar 26,Apr 26,May 26'. "
                             "Defaults to the last 3 calendar months.")
    parser.add_argument("--title", default="Cloud Security — Governance Report",
                        help="Title shown on the Summary tab.")
    parser.add_argument("--org",   default="PROPRIETARY",
                        help="Organisation name used in the confidentiality watermark.")
    args = parser.parse_args()

    # Collect missing inputs
    if not args.mode or (args.mode != "remove" and not args.input):
        args = _prompt_inputs(args)
    elif not args.mode:
        if Path(args.output).is_file():
            choice = input(
                f"'{args.output}' already exists.\n"
                "  (A) Append / update tabs   (N) Create new file\n"
                "Choice [A]: "
            ).strip().upper() or "A"
            args.mode = "append" if choice.startswith("A") else "new"
        else:
            args.mode = "new"

    # ── Remove mode ────────────────────────────────────────────────────────────
    if args.mode == "remove":
        if not Path(args.output).is_file():
            print(f"Error: workbook not found: {args.output}")
            sys.exit(1)

        to_remove = [s.strip().upper() for s in (args.remove or [])]
        if not to_remove:
            print("Error: no severities specified for removal. Use --remove or the interactive prompt.")
            sys.exit(1)

        print(f"\nLoading {args.output} ...")
        wb = load_workbook(args.output)

        for sev in to_remove:
            tab_name = f"{sev} Issues"
            if tab_name in wb.sheetnames:
                del wb[tab_name]
                print(f"  Removed '{tab_name}'.")
            else:
                print(f"  Tab '{tab_name}' not found — skipped.")

        # Read metrics from all remaining severity tabs
        severity_metrics: dict[str, dict] = {}
        for sev in SEVERITY_ORDER:
            existing = read_metrics_from_existing_tab(wb, sev)
            if existing:
                severity_metrics[sev] = existing

        if severity_metrics:
            # Reconstruct fake snapshots (labels only) from the first remaining tab
            snap_labels = next(iter(severity_metrics.values()))["labels"]
            fake_snapshots = [(lbl, datetime(2000, 1, 1)) for lbl in snap_labels]
            print("  Rebuilding Summary tab ...")
            build_summary_tab(wb, severity_metrics, fake_snapshots, title=args.title)
        else:
            if "Summary" in wb.sheetnames:
                del wb["Summary"]
                print("  No severity tabs remain — Summary removed.")

        apply_watermark(wb, org=args.org)
        while True:
            try:
                wb.save(args.output)
                print(f"\nDone. Report saved → {Path(args.output).resolve()}")
                break
            except PermissionError:
                input(
                    f"\n  Cannot save '{args.output}' — the file is open in Excel.\n"
                    "  Close it and press Enter to retry: "
                )
        return

    # ── New / Append mode ──────────────────────────────────────────────────────

    # Resolve snapshots
    snapshots = (
        parse_snapshots(args.snapshots) if args.snapshots else default_snapshots()
    )
    if not snapshots:
        print("Error: no valid snapshot dates.")
        sys.exit(1)

    # Load CSV(s) — either pre-loaded (inline paste) or from files
    all_rows: list[dict] = []
    pre = getattr(args, '_pre_loaded_rows', None)
    if pre:
        all_rows = pre
        print(f"\n  {len(all_rows)} issues loaded.")
    else:
        for csv_path in args.input:
            print(f"\nLoading {csv_path} ...")
            file_rows = load_wiz_csv(csv_path)
            all_rows.extend(file_rows)
            print(f"  {len(file_rows)} issues loaded.")
        if len(args.input) > 1:
            print(f"  Total: {len(all_rows)} issues across {len(args.input)} files.")

    # Group by severity
    by_severity: dict[str, list[dict]] = defaultdict(list)
    for r in all_rows:
        sev = r.get("Severity", "").strip().upper()
        if sev:
            by_severity[sev].append(r)

    found = sorted(by_severity, key=lambda s: SEVERITY_ORDER.index(s) if s in SEVERITY_ORDER else 99)
    print(f"  Severities detected : {', '.join(found)}")
    print(f"  Snapshot dates      : {', '.join(s[0] for s in snapshots)}")
    print(f"  Mode                : {args.mode}")

    # Open or create workbook
    if args.mode == "append" and Path(args.output).is_file():
        wb = load_workbook(args.output)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Build severity tabs from CSV data
    severity_metrics: dict[str, dict] = {}

    # In append mode: first harvest metrics from existing tabs for severities
    # NOT being replaced — so they still appear correctly in the Summary tab.
    if args.mode == "append" and Path(args.output).is_file():
        for sev in SEVERITY_ORDER:
            if sev not in by_severity:
                existing = read_metrics_from_existing_tab(wb, sev, len(snapshots))
                if existing:
                    severity_metrics[sev] = existing
                    print(f"  Preserving {sev} metrics from existing tab.")

    for sev in SEVERITY_ORDER:
        if sev not in by_severity:
            continue
        rows = by_severity[sev]
        print(f"  Building {sev} tab ({len(rows)} issues) ...")
        build_severity_tab(wb, sev, rows, snapshots)
        severity_metrics[sev] = compute_metrics(rows, snapshots)

    # Summary tab (always first)
    print("  Building Summary tab ...")
    build_summary_tab(wb, severity_metrics, snapshots, title=args.title)

    # Watermark
    apply_watermark(wb, org=args.org)

    while True:
        try:
            wb.save(args.output)
            print(f"\nDone. Report saved → {Path(args.output).resolve()}")
            break
        except PermissionError:
            input(
                f"\n  Cannot save '{args.output}' — the file is open in Excel.\n"
                "  Close it and press Enter to retry: "
            )


if __name__ == "__main__":
    main()
