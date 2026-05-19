import argparse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(description="Convert a CSV into a single styled Excel sheet, dropping specified columns.")
parser.add_argument("csv_file", help="Path to the input CSV file")
parser.add_argument("output", help="Path for the output .xlsx file")
parser.add_argument("--drop", nargs="*", default=["Subscription"], metavar="COL",
                    help="Column name(s) to exclude (default: Subscription)")
args = parser.parse_args()

drop = set(args.drop)

with open(args.csv_file, newline="") as f:
    reader = csv.reader(f)
    all_rows = list(reader)

headers = all_rows[0]
rows = all_rows[1:]
keep_idx = [i for i, h in enumerate(headers) if h not in drop]
new_headers = ["S.No"] + [headers[i] for i in keep_idx]

wb = Workbook()
ws = wb.active
ws.title = "AFFECTED VMs"

for ci, h in enumerate(new_headers, 1):
    ws.cell(row=1, column=ci, value=h)

for ri, row in enumerate(rows, 1):
    ws.cell(row=ri + 1, column=1, value=ri)
    for ci, idx in enumerate(keep_idx, 2):
        ws.cell(row=ri + 1, column=ci, value=row[idx])

# Styling
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
thin = Side(style="thin", color="CCCCCC")
border = Border(bottom=thin)
light = PatternFill("solid", fgColor="F2F7FB")

for ci in range(1, len(new_headers) + 1):
    cell = ws.cell(row=1, column=ci)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.freeze_panes = "A2"

for row_idx in range(2, len(rows) + 2):
    for cell in ws[row_idx]:
        cell.border = border
        if (row_idx - 2) % 2 == 0:
            cell.fill = light

for col_cells in ws.columns:
    lengths = [len(str(c.value)) if c.value else 0 for c in col_cells]
    best = max(lengths) + 2
    col_letter = get_column_letter(col_cells[0].column)
    ws.column_dimensions[col_letter].width = max(10, min(best, 50))

wb.save(args.output)
print(f"Done. {len(rows)} rows, columns: {new_headers}")
