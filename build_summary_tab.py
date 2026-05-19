import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(description="Add/replace the Summary tab in a governance workbook.")
parser.add_argument("workbook", help="Path to the .xlsx workbook")
args = parser.parse_args()

# ── Styles ────────────────────────────────────────────────────────────────────
thin      = Side(style='thin', color='BDD7EE')
def cell_border(): return Border(top=thin, bottom=thin, left=thin, right=thin)

HDR_DARK   = PatternFill('solid', fgColor='1F4E79')
HDR_GREEN  = PatternFill('solid', fgColor='375623')
CRIT_FILL  = PatternFill('solid', fgColor='C00000')
HIGH_FILL  = PatternFill('solid', fgColor='C55A11')
ZEBRA      = PatternFill('solid', fgColor='EBF3FB')
GOLD_FILL  = PatternFill('solid', fgColor='FFF2CC')

CTR_WRAP   = Alignment(horizontal='center', vertical='center', wrap_text=True)
CTR        = Alignment(horizontal='center', vertical='center')
LEFT_PAD   = Alignment(horizontal='left',   vertical='center', indent=1)

WHITE_BOLD = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
DARK_BOLD  = Font(name='Calibri', bold=True, size=11, color='1F4E79')
REG        = Font(name='Calibri', size=11)

def h(ws, row, col, val, fill=None, font=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=val)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c

# ── Data (from individual tabs) ───────────────────────────────────────────────
# Columns: Severity | Total Issues | Open (May 15) | Resolved | Ignored | Resolution % | vs Baseline
summary_data = [
    ('CRITICAL', 'C00000', 8,   4,  3,  1,  '50.0%', '+50.0%'),
    ('HIGH',     'C55A11', 112, 30, 65, 17, '73.2%', '+73.2%'),
]

# ── Build tab ─────────────────────────────────────────────────────────────────
wb = load_workbook(args.workbook)

if 'Summary' in wb.sheetnames:
    del wb['Summary']

ws = wb.create_sheet('Summary', 0)   # insert as first tab
ws.sheet_properties.tabColor = '1F4E79'

# ── Title ─────────────────────────────────────────────────────────────────────
ws.merge_cells('A1:H1')
h(ws, 1, 1, 'CareFirst Cloud Security — Governance Report',
  fill=HDR_DARK,
  font=Font(name='Calibri', bold=True, size=16, color='FFFFFF'),
  align=CTR_WRAP)
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:H2')
h(ws, 2, 1, 'Exec Summary  |  Snapshot: May 15, 2026  |  All Severities',
  fill=PatternFill('solid', fgColor='3C3C3C'),
  font=Font(name='Calibri', italic=True, size=10, color='FFFFFF'),
  align=LEFT_PAD)
ws.row_dimensions[2].height = 18

# ── Section label ─────────────────────────────────────────────────────────────
ws.merge_cells('A3:D3')
h(ws, 3, 1, 'LATEST SNAPSHOT  (May 15, 2026)',
  fill=PatternFill('solid', fgColor='2E75B6'), font=WHITE_BOLD, align=CTR)
for ci in range(2, 5):
    ws.cell(row=3, column=ci).fill = PatternFill('solid', fgColor='2E75B6')
ws.merge_cells('E3:H3')
h(ws, 3, 5, 'PROGRESS vs BASELINE  (Mar 15, 2026)',
  fill=HDR_GREEN, font=WHITE_BOLD, align=CTR)
for ci in range(6, 9):
    ws.cell(row=3, column=ci).fill = HDR_GREEN
ws.row_dimensions[3].height = 22

# ── Column headers ────────────────────────────────────────────────────────────
col_headers = [
    'Severity',
    'Total\nIssues',
    'Open /\nIn Progress',
    'Resolved\n(Cumulative)',
    'Ignored /\nAccepted Risk',
    'Resolution %\n(Resolved + Ignored)',
    'vs Baseline\n(Mar → May)',
    'Trend',
]
for ci, ch in enumerate(col_headers, 1):
    fill = HDR_DARK if ci <= 5 else HDR_GREEN
    h(ws, 4, ci, ch, fill=fill, font=WHITE_BOLD, align=CTR_WRAP, border=cell_border())
ws.row_dimensions[4].height = 44

# ── Data rows ─────────────────────────────────────────────────────────────────
trend_map = {
    'CRITICAL': '↓ Open issues declining',
    'HIGH':     '↓ Open issues declining',
}
for ri, (sev, color, total, open_, res, ign, pct, vs_base) in enumerate(summary_data, 5):
    sev_fill = PatternFill('solid', fgColor=color)
    row_fill = ZEBRA if ri % 2 == 0 else None
    vals = [sev, total, open_, res, ign, pct, vs_base, trend_map.get(sev, '')]
    for ci, val in enumerate(vals, 1):
        if ci == 1:
            h(ws, ri, ci, val, fill=sev_fill,
              font=Font(name='Calibri', bold=True, size=12, color='FFFFFF'),
              align=CTR, border=cell_border())
        elif ci == 6:   # Resolution %
            h(ws, ri, ci, val, fill=GOLD_FILL,
              font=Font(name='Calibri', bold=True, size=12, color='7F6000'),
              align=CTR, border=cell_border())
        elif ci == 7:   # vs Baseline
            h(ws, ri, ci, val, fill=HDR_GREEN,
              font=Font(name='Calibri', bold=True, size=12, color='FFFFFF'),
              align=CTR, border=cell_border())
        elif ci == 3:   # Open — highlight red-ish
            h(ws, ri, ci, val, fill=PatternFill('solid', fgColor='FCE4D6'),
              font=Font(name='Calibri', bold=True, size=12, color='C00000'),
              align=CTR, border=cell_border())
        elif ci == 8:   # Trend
            h(ws, ri, ci, val, fill=PatternFill('solid', fgColor='E2EFDA'),
              font=Font(name='Calibri', italic=True, size=10, color='375623'),
              align=CTR, border=cell_border())
        else:
            h(ws, ri, ci, val, fill=row_fill,
              font=Font(name='Calibri', bold=True, size=12),
              align=CTR, border=cell_border())
    ws.row_dimensions[ri].height = 28

# ── Footnotes ─────────────────────────────────────────────────────────────────
note_row = 5 + len(summary_data) + 1
ws.merge_cells(f'A{note_row}:H{note_row}')
h(ws, note_row, 1,
  '* Snapshots taken as of End-of-Day (23:59:59) on each date.  |  See individual severity tabs for month-on-month breakdown.',
  font=Font(name='Calibri', italic=True, size=9, color='808080'),
  align=LEFT_PAD)
ws.row_dimensions[note_row].height = 16

# ── Column widths ─────────────────────────────────────────────────────────────
for ci, w in enumerate([14, 10, 14, 14, 16, 24, 18, 26], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'A5'

wb.save(args.workbook)
print('Done — Summary tab added as first tab.')
